"""
Сервис экспорта данных.

Возможности:
    * экспорт произвольных табличных данных в Excel (pandas + openpyxl)
      с автоподбором ширины колонок и форматированием заголовка;
    * формирование отчёта по себестоимости банкета в Excel;
    * сохранение заявки на закупку в PDF (reportlab) с поддержкой кириллицы.
"""

import os

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
)

from core import config

# Фирменные цвета оформления (синий / серый).
HEADER_FILL = "2F5597"   # синий заголовок таблиц Excel
HEADER_FONT = "FFFFFF"   # белый текст заголовка


# ---------------------------------------------------------------------------
# Регистрация шрифта с поддержкой кириллицы для PDF
# ---------------------------------------------------------------------------

_PDF_FONT = "Helvetica"  # запасной вариант (без кириллицы)


def _register_cyrillic_font() -> str:
    """Зарегистрировать TrueType-шрифт с кириллицей для отчётов PDF.

    Ищет системные шрифты Windows (Arial / DejaVu). Возвращает имя
    зарегистрированного шрифта либо запасной Helvetica.
    """
    global _PDF_FONT
    if _PDF_FONT != "Helvetica":
        return _PDF_FONT  # уже зарегистрирован

    candidates = [
        ("Arial", r"C:\Windows\Fonts\arial.ttf"),
        ("DejaVuSans", r"C:\Windows\Fonts\DejaVuSans.ttf"),
        ("Calibri", r"C:\Windows\Fonts\calibri.ttf"),
    ]
    for name, path in candidates:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont(name, path))
                _PDF_FONT = name
                return name
            except Exception:
                continue
    return _PDF_FONT


# ---------------------------------------------------------------------------
# Экспорт в Excel
# ---------------------------------------------------------------------------

def export_dataframe_to_excel(df: pd.DataFrame, path: str,
                              sheet_name: str = "Отчёт",
                              title: str = "") -> None:
    """Сохранить DataFrame в Excel с оформлением заголовка и автошириной.

    :param df:    данные для выгрузки.
    :param path:  путь к создаваемому файлу .xlsx.
    :param title: необязательный заголовок над таблицей.
    """
    start_row = 1 if title else 0
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False,
                    startrow=start_row)
        ws = writer.sheets[sheet_name]

        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        if title:
            ws.cell(row=1, column=1, value=title)
            cell = ws.cell(row=1, column=1)
            cell.font = Font(bold=True, size=14, color="1F3864")
            ws.merge_cells(start_row=1, start_column=1,
                           end_row=1, end_column=max(1, len(df.columns)))

        # Оформление строки заголовков таблицы.
        header_row = start_row + 1
        for col_idx, column in enumerate(df.columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = Font(bold=True, color=HEADER_FONT)
            cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border = border

        # Границы для строк данных.
        for r in range(header_row + 1, header_row + 1 + len(df)):
            for c in range(1, len(df.columns) + 1):
                ws.cell(row=r, column=c).border = border

        # Автоподбор ширины колонок по содержимому.
        for col_idx, column in enumerate(df.columns, start=1):
            max_len = len(str(column))
            for value in df.iloc[:, col_idx - 1]:
                max_len = max(max_len, len(str(value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = \
                min(max_len + 4, 50)


def export_banquet_cost_to_excel(cost: dict, path: str) -> None:
    """Сохранить отчёт по себестоимости банкета в Excel.

    :param cost: результат services.costing.calc_banquet_cost.
    """
    banquet = cost["banquet"]

    # Детализация по ингредиентам.
    rows = []
    for item in cost["items"]:
        rows.append({
            "Ингредиент": item["name"],
            "Ед. изм.": item["unit"],
            "Цена, ₽": round(item["price"], 2),
            "Требуется": round(item["required"], 3),
            "Стоимость, ₽": round(item["cost"], 2),
        })
    df = pd.DataFrame(rows)

    # Сводный блок.
    summary = pd.DataFrame({
        "Показатель": [
            "Клиент", "Дата", "Гостей",
            "Стоимость продуктов, ₽",
            f"Накладные расходы ({cost['overhead_pct']:.0f}%), ₽",
            "ИТОГО себестоимость, ₽", "Себестоимость на гостя, ₽",
        ],
        "Значение": [
            banquet["client_name"], banquet["event_date"], cost["guests"],
            round(cost["ingredients_cost"], 2), round(cost["overhead"], 2),
            round(cost["total"], 2), round(cost["cost_per_guest"], 2),
        ],
    })

    title = f"Калькуляция себестоимости банкета: {banquet['client_name']}"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Себестоимость", index=False,
                         startrow=1)
        ws = writer.sheets["Себестоимость"]
        ws.cell(row=1, column=1, value=title).font = \
            Font(bold=True, size=14, color="1F3864")
        ws.merge_cells("A1:B1")
        for c in range(1, 3):
            cell = ws.cell(row=2, column=c)
            cell.font = Font(bold=True, color=HEADER_FONT)
            cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 26

        if not df.empty:
            df.to_excel(writer, sheet_name="Расход продуктов", index=False)
            ws2 = writer.sheets["Расход продуктов"]
            for c in range(1, len(df.columns) + 1):
                cell = ws2.cell(row=1, column=c)
                cell.font = Font(bold=True, color=HEADER_FONT)
                cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
                cell.alignment = Alignment(horizontal="center")
            for col_idx in range(1, len(df.columns) + 1):
                ws2.column_dimensions[get_column_letter(col_idx)].width = 18


# ---------------------------------------------------------------------------
# Экспорт заявки на закупку в PDF
# ---------------------------------------------------------------------------

def export_purchase_to_pdf(purchase: dict, items: list, path: str) -> None:
    """Сохранить заявку на закупку в файл PDF.

    :param purchase: запись закупки (sqlite3.Row или dict).
    :param items:    позиции закупки.
    :param path:     путь к создаваемому файлу .pdf.
    """
    font = _register_cyrillic_font()
    bold_font = font  # для TTF используем тот же шрифт (начертание имитируется)

    doc = SimpleDocTemplate(
        path, pagesize=A4,
        leftMargin=20 * mm, rightMargin=20 * mm,
        topMargin=18 * mm, bottomMargin=18 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleRu", parent=styles["Title"], fontName=font, fontSize=16,
        textColor=colors.HexColor("#1F3864"),
    )
    normal = ParagraphStyle(
        "NormalRu", parent=styles["Normal"], fontName=font, fontSize=10,
        leading=14,
    )

    org = config.get_setting("organization", "ООО «Агнес»")
    city = config.get_setting("city", "Санкт-Петербург")

    elements = [
        Paragraph(f"{org}, {city}", normal),
        Spacer(1, 4 * mm),
        Paragraph(f"ЗАЯВКА НА ЗАКУПКУ № {purchase['id']}", title_style),
        Spacer(1, 4 * mm),
        Paragraph(f"Дата формирования: {purchase['created_at']}", normal),
        Paragraph(f"Поставщик: {purchase['supplier_name'] or '—'}", normal),
        Paragraph(f"ИНН поставщика: {purchase['supplier_inn'] or '—'}", normal),
        Paragraph(f"Статус: {purchase['status']}", normal),
        Spacer(1, 6 * mm),
    ]

    # Таблица позиций.
    data = [["№", "Наименование", "Ед.", "Кол-во", "Цена, ₽", "Сумма, ₽"]]
    total = 0.0
    for idx, item in enumerate(items, start=1):
        line_sum = (item["qty"] or 0) * (item["price"] or 0)
        total += line_sum
        data.append([
            str(idx), item["ingredient_name"], item["unit"],
            f"{item['qty']:.3f}", f"{item['price']:.2f}", f"{line_sum:.2f}",
        ])
    data.append(["", "", "", "", "ИТОГО:", f"{total:.2f}"])

    table = Table(data, colWidths=[12 * mm, 65 * mm, 15 * mm,
                                   22 * mm, 25 * mm, 28 * mm])
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5597")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("ALIGN", (2, 1), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BFBFBF")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2),
         [colors.white, colors.HexColor("#F2F5FB")]),
        ("FONTNAME", (4, -1), (-1, -1), bold_font),
        ("SPAN", (0, -1), (4, -1)),
        ("ALIGN", (0, -1), (4, -1), "RIGHT"),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 12 * mm))
    elements.append(Paragraph("Ответственный менеджер: "
                              "_______________ / ________________ /", normal))

    doc.build(elements)
